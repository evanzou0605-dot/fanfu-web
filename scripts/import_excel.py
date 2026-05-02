from __future__ import annotations

import json
import re
from copy import deepcopy
from datetime import date, datetime
from hashlib import sha1
from pathlib import Path

ROOT = Path("/Users/jianweizou/Documents/codex/fanfu_web")
WORKBOOK_PATH = Path("/Users/jianweizou/anticorruption-site/二十大打虎记—ai编辑版.xlsx")
OFFICIALS_PATH = ROOT / "data" / "officials.json"

REGIONS = {
    "北京市",
    "天津市",
    "河北省",
    "山西省",
    "内蒙古自治区",
    "辽宁省",
    "吉林省",
    "黑龙江省",
    "上海市",
    "江苏省",
    "浙江省",
    "安徽省",
    "福建省",
    "江西省",
    "山东省",
    "河南省",
    "湖北省",
    "湖南省",
    "广东省",
    "广西壮族自治区",
    "海南省",
    "重庆市",
    "四川省",
    "贵州省",
    "云南省",
    "西藏自治区",
    "陕西省",
    "甘肃省",
    "青海省",
    "宁夏回族自治区",
    "新疆维吾尔自治区",
    "中央部委/央企",
}

SHEET_REGION_MAP = {
    "西藏藏族自治区": "西藏自治区",
    "中央部委": "中央部委/央企",
}

SKIP_SHEETS = {"Sheet1", "军工系", "医疗反腐"}

LEVELS = {"国家级", "省部级", "厅局级"}

REGION_ALIASES = {
    "北京": "北京市",
    "天津": "天津市",
    "河北": "河北省",
    "山西": "山西省",
    "内蒙古": "内蒙古自治区",
    "辽宁": "辽宁省",
    "吉林": "吉林省",
    "黑龙江": "黑龙江省",
    "上海": "上海市",
    "江苏": "江苏省",
    "浙江": "浙江省",
    "安徽": "安徽省",
    "福建": "福建省",
    "江西": "江西省",
    "山东": "山东省",
    "河南": "河南省",
    "湖北": "湖北省",
    "湖南": "湖南省",
    "广东": "广东省",
    "广西": "广西壮族自治区",
    "海南": "海南省",
    "重庆": "重庆市",
    "四川": "四川省",
    "贵州": "贵州省",
    "云南": "云南省",
    "西藏": "西藏自治区",
    "陕西": "陕西省",
    "甘肃": "甘肃省",
    "青海": "青海省",
    "宁夏": "宁夏回族自治区",
    "新疆": "新疆维吾尔自治区",
}

KNOWN_LOCATION_PREFIXES = sorted(
    {**REGION_ALIASES, **{region: region for region in REGIONS if region != "中央部委/央企"}}.items(),
    key=lambda item: len(item[0]),
    reverse=True,
)


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").replace("\t", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip(" ,，;；")


def normalize_date(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    text = normalize_text(value)
    if not text:
        return ""
    match = re.search(r"(\d{4})[./年-](\d{1,2})[./月-](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"
    match = re.search(r"(\d{4})[./年-](\d{1,2})", text)
    if match:
        return f"{match.group(1)}-{int(match.group(2)):02d}-01"
    return text


def infer_region(sheet_name: str, positions: str) -> str:
    mapped = SHEET_REGION_MAP.get(sheet_name, sheet_name)
    if mapped in REGIONS:
        return mapped
    source = normalize_text(positions)
    for alias, region in REGION_ALIASES.items():
        if alias in source or region in source:
            return region
    return "中央部委/央企"


def infer_level_from_text(text: str) -> str:
    input_text = normalize_text(text)
    cleaned = (
        input_text.replace("中央纪委国家监委", " ")
        .replace("中央纪委", " ")
        .replace("国家监委", " ")
        .replace("国家监察委员会", " ")
    )
    if "中央政治局" in cleaned or "国务委员" in cleaned:
        return "国家级"
    if re.search(r"(北京市|上海市|天津市|重庆市|北京|上海|天津|重庆).{0,12}市委常委", cleaned):
        return "省部级"
    if re.search(r"(北京市|上海市|天津市|重庆市|北京|上海|天津|重庆).{0,12}(副市长|市政协主席|市人大常委会主任)", cleaned):
        return "省部级"
    if any(
        token in cleaned
        for token in [
            "省长",
            "副省长",
            "省委书记",
            "省委常委",
            "省政协",
            "省人大",
            "部长",
            "副部长",
            "党组书记",
            "董事长",
        ]
    ):
        return "省部级"
    return "厅局级"


LOCATION_PATTERN = re.compile(r"^([\u4e00-\u9fa5]{2,12}?(?:自治区|自治州|新区|省|市|区|县|州|盟|旗))")
TRAILING_PARTY_TITLES = ("党组书记", "党组副书记", "党委书记", "党委副书记")
GENERIC_LOCATIONS = {"省", "市", "区", "县", "州", "盟", "旗", "自治区", "自治州", "新区"}
LOCATION_SUFFIXES = ("自治区", "自治州", "新区", "省", "市", "区", "县", "州", "盟", "旗")
PROVINCE_LEVEL_PREFIXES = (
    "省委",
    "省政府",
    "省人民政府",
    "省人大",
    "省政协",
    "自治区党委",
    "自治区政府",
    "自治区人民政府",
    "自治区人大",
    "自治区政协",
)
CITY_LEVEL_PREFIXES = (
    "市委",
    "市政府",
    "市人民政府",
    "市人大",
    "市政协",
    "州委",
    "州政府",
    "州人民政府",
    "区委",
    "区政府",
    "区人民政府",
    "县委",
    "县政府",
    "县人民政府",
)
TERMINAL_AUXILIARY_PATTERNS = (
    r"选举委员会委员",
    r"工作协调小组成员",
    r"领导小组成员",
    r"委员会委员$",
)


def extract_location_context(text: str) -> str:
    normalized = normalize_text(text)
    for prefix, mapped in KNOWN_LOCATION_PREFIXES:
        if normalized.startswith(prefix):
            return mapped

    matches = LOCATION_PATTERN.findall(normalize_text(text))
    if not matches:
        return ""
    location = matches[-1]
    if any(token in location for token in ["政府", "党委", "人大", "政协", "纪委", "监察", "法院", "检察院", "党校"]):
        return ""
    return "" if location in GENERIC_LOCATIONS else location


def normalize_region_prefix(region: str, part: str) -> str:
    if not region:
        return ""
    normalized_part = normalize_text(part)
    if region.endswith("省"):
        return region
    if region.endswith("自治区"):
        return region if normalized_part.startswith("自治区") else region
    if region.endswith("市"):
        return region
    return region


def join_inherited_location(base_location: str, part: str) -> str:
    if not base_location:
        return normalize_text(part)
    inherited = f"{base_location}{normalize_text(part)}"
    for suffix in LOCATION_SUFFIXES:
        if base_location.endswith(suffix) and normalize_text(part).startswith(suffix):
            inherited = f"{base_location}{normalize_text(part)[len(suffix):]}"
            break
    return inherited


def is_province_level_fragment(part: str) -> bool:
    normalized = normalize_text(part)
    return any(normalized.startswith(prefix) for prefix in PROVINCE_LEVEL_PREFIXES)


def is_city_level_fragment(part: str) -> bool:
    normalized = normalize_text(part)
    return any(normalized.startswith(prefix) for prefix in CITY_LEVEL_PREFIXES)


def inherit_location_context(parts: list[str], region: str = "") -> list[str]:
    normalized = []
    current_location = ""
    province_context = region if region in REGIONS and region != "中央部委/央企" else ""
    for raw in parts:
        part = normalize_text(raw)
        if not part:
            continue
        explicit_location = extract_location_context(part)
        if explicit_location:
            current_location = explicit_location
            if explicit_location.endswith("省") or explicit_location.endswith("自治区") or explicit_location.endswith("市"):
                province_context = province_context or explicit_location
            normalized.append(part)
            continue

        inherited_base = ""
        if is_province_level_fragment(part):
            inherited_base = normalize_region_prefix(province_context, part)
        elif is_city_level_fragment(part):
            inherited_base = current_location

        if inherited_base and re.match(r"^(自治区|自治州|市|市委|市政府|市人民政府|市人大|市政协|市纪委|市监察局|市委党校|市检察院|市法院|省|省委|省政府|省人民政府|省人大|省政协|自治区党委|自治区政府|自治区人民政府|自治区人大|自治区政协|区|区委|区政府|区人民政府|县|县委|县政府|县人民政府|州|州委|州政府|州人民政府)", part):
            inherited = join_inherited_location(inherited_base, part)
            normalized.append(inherited)
        else:
            normalized.append(part)
    return normalized


def merge_trailing_titles(parts: list[str]) -> list[str]:
    if not parts:
        return []
    merged = []
    for part in parts:
        if merged and any(part.startswith(title) for title in TRAILING_PARTY_TITLES):
            merged[-1] = f"{merged[-1]}、{part}"
        else:
            merged.append(part)
    return merged


def same_location(left: str, right: str) -> bool:
    return extract_location_context(left) and extract_location_context(left) == extract_location_context(right)


def merge_terminal_office(parts: list[str]) -> list[str]:
    if len(parts) < 2:
        return parts

    merged = parts[:]
    last = merged[-1]
    prev = merged[-2]
    if same_location(prev, last) and re.search(r"(省委副书记|市委副书记)", prev) and re.search(
        r"(政府.*省长|政府.*市长|政府.*主席|人民政府.*省长|人民政府.*市长|人民政府.*主席)", last
    ):
        merged[-2:] = [f"{prev}，{last}"]
    return merged


def normalize_position_phrase(text: str) -> str:
    value = normalize_text(text)
    location = extract_location_context(value)

    if location:
        value = value.replace(f"{location}副省人大常委会", f"{location}省人大常委会")
        value = value.replace(f"{location}副省政协", f"{location}省政协")
        value = value.replace(f"{location}人民政府副省政协", f"{location}省政协")
        value = value.replace(f"{location}人民政府副市委常委", f"{location}市委常委")
        value = value.replace(f"{location}人民政府副省人大常委会", f"{location}省人大常委会")

    value = re.sub(r"^([\u4e00-\u9fa5]{2,6})省委副书记，\1省人民政府", r"\1省委副书记，省人民政府", value)
    value = re.sub(r"^([\u4e00-\u9fa5]{2,6})市委副书记，\1市人民政府", r"\1市委副书记，市人民政府", value)
    value = value.replace("中国共产党中央军事委员会委员", "中央军委委员")
    value = value.replace("中华人民共和国中央军事委员会委员", "中央军委委员")

    return value


def repair_mismatched_admin_level(text: str, region: str) -> str:
    value = normalize_text(text)
    normalized_region = normalize_text(region)
    if not value or not normalized_region or normalized_region == "中央部委/央企":
        return value

    if normalized_region.endswith("省"):
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(省政协)", f"{normalized_region}政协", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(省人大)", f"{normalized_region}人大", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(省政府)", f"{normalized_region}政府", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(省人民政府)", f"{normalized_region}人民政府", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(省委)", f"{normalized_region}委", value)
    elif normalized_region.endswith("自治区"):
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(自治区政协)", f"{normalized_region}政协", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(自治区人大)", f"{normalized_region}人大", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(自治区政府)", f"{normalized_region}政府", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(自治区人民政府)", f"{normalized_region}人民政府", value)
        value = re.sub(r"^[\u4e00-\u9fa5]{2,12}(?:市|州|盟|区|县)(自治区党委)", f"{normalized_region}党委", value)
    return value


def is_auxiliary_terminal_position(text: str) -> bool:
    value = normalize_text(text)
    return any(re.search(pattern, value) for pattern in TERMINAL_AUXILIARY_PATTERNS)


def combine_central_military_positions(parts: list[str]) -> list[str]:
    normalized = []
    has_cmc_member = False
    has_political_work_director = False
    has_state_councilor = False
    has_defense_minister = False
    has_party_cmc_vice_chair = False
    has_state_cmc_vice_chair = False

    for part in parts:
        value = normalize_text(part)
        if not value:
            continue
        if value == "中央军委委员":
            has_cmc_member = True
            continue
        if value in {"国务委员、国务院党组成员", "国务委员"}:
            has_state_councilor = True
            continue
        if value == "国防部部长":
            has_defense_minister = True
            continue
        if value == "中共中央军事委员会副主席":
            has_party_cmc_vice_chair = True
            continue
        if value == "中华人民共和国中央军事委员会副主席":
            has_state_cmc_vice_chair = True
            continue
        if value == "中央军委政治工作部主任":
            has_political_work_director = True
            continue
        normalized.append(value)

    if has_party_cmc_vice_chair or has_state_cmc_vice_chair:
        titles = []
        if has_party_cmc_vice_chair:
            titles.append("中共中央军事委员会副主席")
        if has_state_cmc_vice_chair:
            titles.append("中华人民共和国中央军事委员会副主席")
        normalized.append("、".join(titles))

    if has_cmc_member or has_state_councilor or has_defense_minister or has_political_work_director:
        titles = []
        if has_cmc_member:
            titles.append("中央军委委员")
        if has_state_councilor:
            titles.append("国务委员")
        if has_defense_minister:
            titles.append("国防部部长")
        if has_political_work_director:
            titles.append("中央军委政治工作部主任")
        normalized.append("、".join(titles))

    return normalized


def choose_last_position(parts: list[str]) -> tuple[str, list[str]]:
    if not parts:
        return "", []

    last_index = len(parts) - 1
    while last_index > 0 and is_auxiliary_terminal_position(parts[last_index]):
        last_index -= 1

    last_position = parts[last_index]
    previous_positions = [part for index, part in enumerate(parts) if index != last_index]
    return last_position, previous_positions


def build_positions(raw: str, region: str = "") -> tuple[str, list[str]]:
    text = normalize_text(raw)
    if not text:
        return "", []
    parts = [normalize_text(item) for item in re.split(r"[，,；;]", text) if normalize_text(item)]
    parts = inherit_location_context(parts, region)
    parts = merge_trailing_titles(parts)
    parts = merge_terminal_office(parts)
    parts = [normalize_position_phrase(part) for part in parts]
    parts = [repair_mismatched_admin_level(part, region) for part in parts]
    parts = combine_central_military_positions(parts)
    if not parts:
        return text, []
    if len(parts) == 1:
        return parts[0], []
    last_position, previous_positions = choose_last_position(parts)
    previous_positions = previous_positions[-5:]
    return last_position, previous_positions


def make_id(name: str, investigation_date: str) -> str:
    return sha1(f"{name}|{investigation_date}".encode("utf-8")).hexdigest()[:16]


def build_summary(name: str, date_text: str, result_text: str) -> str:
    if result_text:
        return f"{name}于{date_text or '未知时间'}被通报审查调查，后续处理结果为：{result_text}"
    return f"{name}于{date_text or '未知时间'}被通报审查调查"


def build_detail(name: str, positions: str, result_text: str, sheet_name: str) -> str:
    parts = [f"来自 Excel 工作表《{sheet_name}》的历史整理数据。"]
    if positions:
        parts.append(f"曾任职务：{positions}")
    if result_text:
        parts.append(f"处理结果：{result_text}")
    return " ".join(parts)


def build_timeline(investigation_date: str, summary: str, result_text: str) -> list[dict]:
    timeline = []
    if investigation_date or summary:
        timeline.append(
            {
                "stage": "审查调查",
                "date": investigation_date,
                "url": "",
                "summary": summary,
            }
        )
    if result_text:
        timeline.append(
            {
                "stage": "后续处理",
                "date": "",
                "url": "",
                "summary": result_text,
            }
        )
    return timeline


def merge_timeline(existing: list[dict], incoming: list[dict]) -> list[dict]:
    seen = {}
    for item in [*(existing or []), *(incoming or [])]:
        key = f"{item.get('stage','')}|{item.get('date','')}|{item.get('summary','')}"
        seen[key] = item
    return list(seen.values())


def choose_status(result_text: str) -> str:
    text = normalize_text(result_text)
    return text or "审查调查"


def parse_workbook():
    from openpyxl import load_workbook

    workbook = load_workbook(WORKBOOK_PATH, data_only=True)
    parsed = []
    for sheet_name in workbook.sheetnames:
      if sheet_name in SKIP_SHEETS:
          continue
      ws = workbook[sheet_name]
      current_level = ""
      header_seen = False
      for row in ws.iter_rows(values_only=True):
          cells = [normalize_text(value) for value in row[:5]]
          while len(cells) < 5:
              cells.append("")
          first = cells[0]
          if not any(cells):
              continue
          if first in LEVELS:
              current_level = first
              header_seen = False
              continue
          if first == "姓名":
              if "被调查时间" not in cells[:4]:
                  header_seen = False
                  continue
              header_seen = True
              continue
          if not header_seen:
              continue
          name = first
          if not name or name == "姓名":
              continue
          investigation_date = normalize_date(cells[1])
          positions = cells[2]
          result_text = cells[3]
          region = infer_region(sheet_name, positions)
          last_position, previous_positions = build_positions(positions, region)
          full_positions_text = "，".join([*previous_positions, last_position]) if last_position else positions
          level = current_level or infer_level_from_text(f"{region} {positions}")
          summary = build_summary(name, investigation_date, result_text)
          official = {
              "id": make_id(name, investigation_date),
              "name": name,
              "birth": "",
              "region": region,
              "level": level if level in LEVELS else infer_level_from_text(f"{region} {positions}"),
              "lastPosition": last_position,
              "previousPositions": previous_positions,
              "investigationDate": investigation_date,
              "status": choose_status(result_text),
              "summary": summary,
              "detail": build_detail(name, full_positions_text, result_text, sheet_name),
              "timeline": build_timeline(investigation_date, summary, result_text),
              "sources": [
                  {
                      "type": "excel",
                      "label": f"Excel导入：{sheet_name}",
                      "url": str(WORKBOOK_PATH),
                  }
              ],
              "aliases": [],
              "editable": True,
              "createdAt": datetime.now().isoformat(),
              "updatedAt": datetime.now().isoformat(),
          }
          parsed.append(official)
    return parsed


def merge_existing(parsed: list[dict]) -> list[dict]:
    existing = json.loads(OFFICIALS_PATH.read_text("utf-8")) if OFFICIALS_PATH.exists() else []
    skipped_labels = {f"Excel导入：{sheet}" for sheet in SKIP_SHEETS}
    merged: dict[tuple[str, str], dict] = {}

    for item in existing:
        sources = item.get("sources", [])
        source_labels = {source.get("label", "") for source in sources}
        if source_labels and source_labels.issubset(skipped_labels):
            continue
        item["sources"] = [source for source in sources if source.get("label", "") not in skipped_labels]
        key = (item.get("name", ""), item.get("investigationDate", ""))
        merged[key] = deepcopy(item)

    for item in parsed:
        key = (item["name"], item["investigationDate"])
        if key not in merged:
            merged[key] = item
            continue
        current = merged[key]
        current_sources = current.get("sources", [])
        is_excel_only = bool(current_sources) and all(source.get("type") == "excel" for source in current_sources)
        current["region"] = item["region"] or current.get("region", "")
        current["level"] = item["level"] or current.get("level", "")
        if is_excel_only or not current.get("lastPosition"):
            current["lastPosition"] = item["lastPosition"]
        if is_excel_only or not current.get("summary"):
            current["summary"] = item["summary"]
        if is_excel_only or not current.get("detail"):
            current["detail"] = item["detail"]
        if is_excel_only or not current.get("status") or current.get("status") == "审查调查":
            current["status"] = item["status"]
        if is_excel_only:
            current["sources"] = item["sources"]
        elif not current.get("sources"):
            current["sources"] = item["sources"]
        else:
            urls = {source.get("url") for source in current.get("sources", [])}
            for source in item["sources"]:
                if source.get("url") not in urls:
                    current["sources"].append(source)
        if is_excel_only or not current.get("previousPositions"):
            current["previousPositions"] = item["previousPositions"]
        current["timeline"] = item["timeline"] if is_excel_only else merge_timeline(current.get("timeline", []), item["timeline"])
        current["updatedAt"] = datetime.now().isoformat()
        merged[key] = current

    result = list(merged.values())
    result.sort(key=lambda item: item.get("investigationDate", ""), reverse=True)
    return result


def main():
    parsed = parse_workbook()
    merged = merge_existing(parsed)
    OFFICIALS_PATH.write_text(json.dumps(merged, ensure_ascii=False, indent=2), "utf-8")
    summary = {
        "parsed": len(parsed),
        "final": len(merged),
        "workbook": str(WORKBOOK_PATH),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
