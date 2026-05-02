from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from hashlib import sha1
from pathlib import Path

from openpyxl import load_workbook

RANKS = {"上将", "中将", "少将"}
DATE_RE = re.compile(r"(\d{4})[./年-](\d{1,2})[./月-](\d{1,2})")


def normalize_text(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).replace("\u3000", " ")).strip(" ,，;；")


def normalize_date(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    match = DATE_RE.search(text)
    if not match:
        return text
    return f"{match.group(1)}-{int(match.group(2)):02d}-{int(match.group(3)):02d}"


def make_id(name: str, investigation_date: str, seed: str) -> str:
    raw = f"{name}|{investigation_date}|{seed}"
    return sha1(raw.encode("utf-8")).hexdigest()[:16]


def split_positions(raw: str) -> list[str]:
    text = normalize_text(raw)
    if not text:
        return []
    parts = [normalize_text(item) for item in re.split(r"[，,；;]", text) if normalize_text(item)]
    merged: list[str] = []
    current = ""
    for part in parts:
        if not current:
            current = part
            continue
        if len(part) <= 8 and any(token in part for token in ["书记", "副书记", "委员", "主任", "政委", "参谋长", "司令员", "副司令员"]):
            current = f"{current}、{part}"
        else:
            merged.append(current)
            current = part
    if current:
        merged.append(current)
    return merged


def infer_last_position(positions: list[str]) -> str:
    return positions[-1] if positions else ""


def build_official(rank: str, name: str, date_text: str, positions_text: str, outcome_text: str, source_label: str) -> dict:
    investigation_date = normalize_date(date_text)
    previous_positions = split_positions(positions_text)
    last_position = infer_last_position(previous_positions)
    summary = (
        f"{name}于{investigation_date}被通报审查调查"
        if investigation_date
        else f"{name} 已录入解放军离线样本，具体落马时间待补"
    )
    detail = f"来自解放军 Excel 离线导入。 曾任职务：{positions_text or last_position or '待补'}"
    timeline = []
    if investigation_date:
        timeline.append(
            {
                "date": investigation_date,
                "stage": "审查调查",
                "summary": summary,
                "url": "",
            }
        )
    if normalize_text(outcome_text):
        timeline.append(
            {
                "date": investigation_date,
                "stage": "后续处理",
                "summary": normalize_text(outcome_text),
                "url": "",
            }
        )
        summary = f"{summary}，后续处理结果为：{normalize_text(outcome_text)}"
        detail = f"{detail} 处理结果：{normalize_text(outcome_text)}"

    return {
        "id": make_id(name, investigation_date, f"{source_label}|{rank}|{positions_text}"),
        "name": name,
        "birth": "",
        "region": "解放军",
        "level": rank,
        "lastPosition": last_position,
        "previousPositions": previous_positions[:-1] if len(previous_positions) > 1 else [],
        "investigationDate": investigation_date,
        "status": "审查调查",
        "summary": summary,
        "detail": detail,
        "timeline": timeline,
        "sources": [
            {
                "type": "excel",
                "label": source_label,
                "url": "",
            }
        ],
        "editable": True,
        "manualRegionOverride": True,
        "manualLevelOverride": True,
        "createdAt": datetime.utcnow().isoformat(),
        "updatedAt": datetime.utcnow().isoformat(),
    }


def parse_workbook(path: Path, source_label: str) -> list[dict]:
    workbook = load_workbook(path, data_only=True)
    officials: list[dict] = []

    for sheet in workbook.worksheets:
        current_rank = ""
        for row in sheet.iter_rows(values_only=True):
            cells = [normalize_text(cell) for cell in row]
            if not any(cells):
                continue
            rank_cell = next((cell for cell in cells if cell in RANKS), "")
            if rank_cell:
                current_rank = rank_cell
                continue
            name = cells[0] if len(cells) > 0 else ""
            if name in RANKS or name in {"姓名", "落马时间", "曾任职务", "处理结果", "最后任职", "军衔"}:
                continue

            date_text = cells[1] if len(cells) > 1 else ""
            positions_text = cells[2] if len(cells) > 2 else ""
            outcome_text = cells[3] if len(cells) > 3 else ""
            row_rank = current_rank

            if len(cells) >= 3 and cells[2] in RANKS:
                positions_text = cells[1]
                row_rank = cells[2]
                date_text = ""
                outcome_text = cells[3] if len(cells) > 3 else ""

            if not row_rank or not name:
                continue
            officials.append(build_official(row_rank, name, date_text, positions_text, outcome_text, source_label))
    return officials


def main() -> int:
    if len(sys.argv) < 2:
      print("Usage: python3 import_military_excel.py <xlsx-path> [source-label]", file=sys.stderr)
      return 1

    workbook_path = Path(sys.argv[1])
    source_label = sys.argv[2] if len(sys.argv) > 2 else f"解放军 Excel 导入：{workbook_path.name}"
    result = parse_workbook(workbook_path, source_label)
    json.dump(result, sys.stdout, ensure_ascii=False)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
